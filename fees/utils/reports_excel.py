import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def export_fees_to_excel(fee_records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fee Records Report"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4f46e5", end_color="4f46e5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Student Name", "Roll No", "Class", "Month", "Due Date", "Status"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows
    for record in fee_records:
        ws.append([
            record.student.name,
            record.student.roll_no,
            record.student.student_class,
            record.month,
            record.due_date.strftime("%Y-%m-%d"),
            record.status
        ])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
